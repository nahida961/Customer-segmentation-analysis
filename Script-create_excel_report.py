"""
create_excel_report.py
Generates a fully formatted Excel workbook:
  • Summary Dashboard
  • Customer Segments
  • Spending Analysis
  • Loyalty Analysis
  • Raw Data
"""

import os
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH   = os.path.join(BASE_DIR, "data", "customers.csv")
DB_PATH     = os.path.join(BASE_DIR, "outputs", "customers.db")
OUTPUT_PATH = os.path.join(BASE_DIR, "outputs", "Customer_Segmentation_Report.xlsx")

os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# ── Load data ──────────────────────────────────────────────────────────────
df = pd.read_csv(DATA_PATH, parse_dates=["join_date", "last_purchase_date"])
conn = sqlite3.connect(DB_PATH)

def q(sql): return pd.read_sql_query(sql, conn)

# ── Style constants ────────────────────────────────────────────────────────
DARK_NAVY   = "1A1A2E"
MID_NAVY    = "0F3460"
ACCENT_RED  = "E94560"
ACCENT_PURP = "533483"
LIGHT_BLUE  = "EBF5FB"
LIGHT_GREY  = "F2F3F4"
WHITE       = "FFFFFF"
DARK_TEXT   = "1C2833"

TIER_COLORS = {
    "Premium":  "1A1A2E",
    "Loyal":    "16213E",
    "Regular":  "0F3460",
    "New":      "533483",
    "Budget":   "E94560",
    "At-Risk":  "F5A623",
}

def hdr_font(sz=11, bold=True, color=WHITE):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def body_font(sz=10, bold=False, color=DARK_TEXT):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left():
    return Alignment(horizontal="left", vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_widths: dict):
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

def style_header_row(ws, row, n_cols, bg=DARK_NAVY, fg=WHITE, sz=11):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = Font(name="Arial", size=sz, bold=True, color=fg)
        cell.fill      = fill(bg)
        cell.alignment = center(wrap=True)
        cell.border    = thin_border()

def style_data_rows(ws, start_row, end_row, n_cols, alt=True):
    for r in range(start_row, end_row + 1):
        bg = LIGHT_BLUE if (alt and r % 2 == 0) else WHITE
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill(bg)
            cell.font      = body_font()
            cell.alignment = left()
            cell.border    = thin_border()

def write_df(ws, df_in, start_row=1, start_col=1,
             header_bg=DARK_NAVY, header_fg=WHITE):
    headers = list(df_in.columns)
    for ci, h in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=ci)
        cell.value     = h.replace("_", " ").title()
        cell.font      = Font(name="Arial", size=10, bold=True, color=header_fg)
        cell.fill      = fill(header_bg)
        cell.alignment = center(wrap=True)
        cell.border    = thin_border()

    for ri, row_data in enumerate(df_in.itertuples(index=False), start=start_row+1):
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row_data, start=start_col):
            cell = ws.cell(row=ri, column=ci)
            cell.value     = None if (isinstance(val, float) and np.isnan(val)) else val
            cell.fill      = fill(bg)
            cell.font      = body_font()
            cell.alignment = left()
            cell.border    = thin_border()

    return start_row + 1 + len(df_in) - 1   # last data row

def kpi_box(ws, row, col, label, formula, fmt="number"):
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font      = Font(name="Arial", size=9, bold=True, color="666666")
    label_cell.fill      = fill(LIGHT_GREY)
    label_cell.alignment = center()
    label_cell.border    = thin_border()

    val_cell = ws.cell(row=row+1, column=col, value=formula)
    val_cell.font      = Font(name="Arial", size=16, bold=True, color=DARK_NAVY)
    val_cell.fill      = fill(WHITE)
    val_cell.alignment = center()
    val_cell.border    = thin_border()
    if fmt == "currency":
        val_cell.number_format = '"$"#,##0'
    elif fmt == "pct":
        val_cell.number_format = "0.0%"
    return val_cell

# ══════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY DASHBOARD
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Summary Dashboard"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 20

# Title banner
ws1.merge_cells("A1:J1")
title = ws1["A1"]
title.value     = "CUSTOMER SEGMENTATION ANALYSIS — EXECUTIVE DASHBOARD"
title.font      = Font(name="Arial", size=16, bold=True, color=WHITE)
title.fill      = fill(DARK_NAVY)
title.alignment = center()

ws1.merge_cells("A2:J2")
sub = ws1["A2"]
sub.value     = "Data Period: Jan 2020 – Dec 2024  |  Total Customers: 500  |  Analysis Date: 2024-12-31"
sub.font      = Font(name="Arial", size=10, italic=True, color=WHITE)
sub.fill      = fill(MID_NAVY)
sub.alignment = center()

# KPI Section header
ws1.row_dimensions[4].height = 22
ws1.merge_cells("A4:J4")
kpi_hdr = ws1["A4"]
kpi_hdr.value     = "KEY PERFORMANCE INDICATORS"
kpi_hdr.font      = Font(name="Arial", size=11, bold=True, color=WHITE)
kpi_hdr.fill      = fill(ACCENT_PURP)
kpi_hdr.alignment = center()

# KPI values (rows 5-6)
ws1.row_dimensions[5].height = 18
ws1.row_dimensions[6].height = 36

kpi_data = q("""
    SELECT
        COUNT(*)                               AS total_customers,
        ROUND(SUM(total_spent),2)              AS total_revenue,
        ROUND(AVG(total_spent),2)              AS avg_clv,
        ROUND(AVG(avg_order_value),2)          AS avg_aov,
        ROUND(AVG(total_purchases),1)          AS avg_purchases,
        SUM(CASE WHEN days_since_last_purchase<=30 THEN 1 ELSE 0 END) AS active_30d
    FROM customers
""").iloc[0]

kpis = [
    ("Total Customers",  kpi_data["total_customers"],  "number"),
    ("Total Revenue",    kpi_data["total_revenue"],    "currency"),
    ("Avg Lifetime Val", kpi_data["avg_clv"],          "currency"),
    ("Avg Order Value",  kpi_data["avg_aov"],          "currency"),
    ("Avg Purchases",    kpi_data["avg_purchases"],    "number"),
    ("Active (30 days)", kpi_data["active_30d"],       "number"),
]
for i, (lbl, val, fmt) in enumerate(kpis, start=2):
    ws1.column_dimensions[get_column_letter(i)].width = 18
    label_cell = ws1.cell(row=5, column=i, value=lbl)
    label_cell.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    label_cell.fill      = fill("2C3E50")
    label_cell.alignment = center()
    label_cell.border    = thin_border()

    val_cell = ws1.cell(row=6, column=i, value=val)
    val_cell.font      = Font(name="Arial", size=16, bold=True, color=DARK_NAVY)
    val_cell.fill      = fill(LIGHT_BLUE)
    val_cell.alignment = center()
    val_cell.border    = thin_border()
    if fmt == "currency":
        val_cell.number_format = '"$"#,##0.00'

# Tier Summary Table
ws1.row_dimensions[9].height = 22
ws1.merge_cells("B9:I9")
tier_hdr = ws1["B9"]
tier_hdr.value     = "CUSTOMER TIER SUMMARY"
tier_hdr.font      = Font(name="Arial", size=11, bold=True, color=WHITE)
tier_hdr.fill      = fill(MID_NAVY)
tier_hdr.alignment = center()

tier_summary = q("""
    SELECT
        customer_tier                                    AS "Tier",
        COUNT(*)                                         AS "Customers",
        ROUND(100.0*COUNT(*)/500,1)                      AS "% of Total",
        ROUND(AVG(total_spent),2)                        AS "Avg CLV ($)",
        ROUND(SUM(total_spent),2)                        AS "Total Revenue ($)",
        ROUND(100.0*SUM(total_spent)/(SELECT SUM(total_spent) FROM customers),1) AS "Revenue %",
        ROUND(AVG(total_purchases),1)                    AS "Avg Purchases",
        ROUND(AVG(loyalty_points),0)                     AS "Avg Loyalty Pts"
    FROM customers
    GROUP BY customer_tier
    ORDER BY AVG(total_spent) DESC
""")

write_df(ws1, tier_summary, start_row=10, start_col=2, header_bg=MID_NAVY)

# Column widths
set_col_width(ws1, {
    "A": 3, "B": 18, "C": 12, "D": 12, "E": 15, "F": 18, "G": 12,
    "H": 16, "I": 18, "J": 3
})

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — CUSTOMER SEGMENTS
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🔍 Customer Segments")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:N1")
ws2["A1"].value     = "CUSTOMER SEGMENTATION — AGE, SPENDING & LOCATION"
ws2["A1"].font      = Font(name="Arial", size=14, bold=True, color=WHITE)
ws2["A1"].fill      = fill(DARK_NAVY)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 32

# ---- Age Group Table ----
ws2.merge_cells("A3:G3")
ws2["A3"].value     = "SEGMENTATION BY AGE GROUP"
ws2["A3"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws2["A3"].fill      = fill(MID_NAVY)
ws2["A3"].alignment = center()

age_df = q("""
    SELECT
        CASE WHEN age BETWEEN 18 AND 24 THEN '18-24 (Gen Z)'
             WHEN age BETWEEN 25 AND 34 THEN '25-34 (Millennials)'
             WHEN age BETWEEN 35 AND 44 THEN '35-44 (Gen X Young)'
             WHEN age BETWEEN 45 AND 54 THEN '45-54 (Gen X Mature)'
             WHEN age BETWEEN 55 AND 64 THEN '55-64 (Boomers)'
             ELSE '65+ (Seniors)' END            AS age_group,
        COUNT(*)                                  AS customers,
        ROUND(AVG(total_spent),2)                 AS avg_total_spent,
        ROUND(AVG(total_purchases),1)             AS avg_purchases,
        ROUND(AVG(avg_order_value),2)             AS avg_order_value,
        ROUND(SUM(total_spent),2)                 AS group_revenue,
        ROUND(100.0*COUNT(*)/500,1)               AS pct_of_total
    FROM customers GROUP BY age_group ORDER BY age_group
""")
write_df(ws2, age_df, start_row=4, start_col=1)

# ---- Spending Band Table ----
r_start = 4 + len(age_df) + 3
ws2.merge_cells(f"A{r_start}:F{r_start}")
ws2[f"A{r_start}"].value     = "SEGMENTATION BY SPENDING BAND"
ws2[f"A{r_start}"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws2[f"A{r_start}"].fill      = fill(ACCENT_PURP)
ws2[f"A{r_start}"].alignment = center()

spend_df = q("""
    SELECT
        CASE WHEN total_spent>=10000 THEN 'Elite (≥$10K)'
             WHEN total_spent>=5000  THEN 'Premium ($5K–$10K)'
             WHEN total_spent>=2000  THEN 'High ($2K–$5K)'
             WHEN total_spent>=500   THEN 'Mid ($500–$2K)'
             WHEN total_spent>=100   THEN 'Low ($100–$500)'
             ELSE 'Minimal (<$100)' END      AS spend_band,
        COUNT(*)                             AS customers,
        ROUND(AVG(total_spent),2)            AS avg_spent,
        ROUND(SUM(total_spent),2)            AS total_revenue,
        ROUND(AVG(total_purchases),1)        AS avg_purchases,
        ROUND(100.0*COUNT(*)/500,1)          AS pct_of_total
    FROM customers GROUP BY spend_band ORDER BY MIN(total_spent) DESC
""")
write_df(ws2, spend_df, start_row=r_start+1, start_col=1, header_bg=ACCENT_PURP)

set_col_width(ws2, {"A":22,"B":12,"C":18,"D":16,"E":16,"F":16,"G":14})

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — SPENDING ANALYSIS
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("💰 Spending Analysis")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:L1")
ws3["A1"].value     = "SPENDING PATTERNS & REVENUE ANALYSIS"
ws3["A1"].font      = Font(name="Arial", size=14, bold=True, color=WHITE)
ws3["A1"].fill      = fill(DARK_NAVY)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 32

# Top 20 High Value Customers
ws3.merge_cells("A3:M3")
ws3["A3"].value     = "TOP 20 HIGH-VALUE CUSTOMERS"
ws3["A3"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3["A3"].fill      = fill(MID_NAVY)
ws3["A3"].alignment = center()

top20 = q("""
    SELECT customer_id, full_name, age, gender, city, state,
           preferred_category,
           ROUND(total_spent,2) AS total_spent,
           total_purchases,
           ROUND(avg_order_value,2) AS avg_order_value,
           loyalty_points,
           days_since_last_purchase,
           customer_tier
    FROM customers ORDER BY total_spent DESC LIMIT 20
""")
last_row = write_df(ws3, top20, start_row=4, start_col=1)

# Highlight top 3
for r in range(5, 8):
    for c in range(1, len(top20.columns)+1):
        ws3.cell(row=r, column=c).fill = fill("FFF9C4")
        ws3.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True, color=DARK_NAVY)

# Cohort Analysis
r2 = last_row + 3
ws3.merge_cells(f"A{r2}:G{r2}")
ws3[f"A{r2}"].value     = "JOIN COHORT ANALYSIS"
ws3[f"A{r2}"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3[f"A{r2}"].fill      = fill(ACCENT_PURP)
ws3[f"A{r2}"].alignment = center()

cohort = q("""
    SELECT STRFTIME('%Y',join_date) AS cohort_year,
           COUNT(*) AS cohort_size,
           ROUND(AVG(total_spent),2) AS avg_clv,
           ROUND(SUM(total_spent),2) AS total_revenue,
           ROUND(AVG(avg_order_value),2) AS avg_aov,
           ROUND(AVG(total_purchases),1) AS avg_purchases
    FROM customers GROUP BY cohort_year ORDER BY cohort_year
""")
write_df(ws3, cohort, start_row=r2+1, start_col=1, header_bg=ACCENT_PURP)

# Revenue by State
r3 = r2 + len(cohort) + 4
ws3.merge_cells(f"A{r3}:F{r3}")
ws3[f"A{r3}"].value     = "TOP 10 STATES BY REVENUE"
ws3[f"A{r3}"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3[f"A{r3}"].fill      = fill(MID_NAVY)
ws3[f"A{r3}"].alignment = center()

state_rev = q("""
    SELECT state, COUNT(*) AS customers,
           ROUND(SUM(total_spent),2) AS total_revenue,
           ROUND(AVG(total_spent),2) AS avg_customer_value,
           ROUND(100.0*SUM(total_spent)/(SELECT SUM(total_spent) FROM customers),1) AS revenue_share_pct
    FROM customers GROUP BY state ORDER BY total_revenue DESC LIMIT 10
""")
write_df(ws3, state_rev, start_row=r3+1, start_col=1)

set_col_width(ws3, {"A":12,"B":22,"C":6,"D":8,"E":12,"F":18,"G":16,
                    "H":16,"I":14,"J":14,"K":14,"L":20,"M":14})

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — LOYALTY ANALYSIS
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🏆 Loyalty Analysis")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:I1")
ws4["A1"].value     = "LOYALTY & RETENTION ANALYSIS"
ws4["A1"].font      = Font(name="Arial", size=14, bold=True, color=WHITE)
ws4["A1"].fill      = fill(DARK_NAVY)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 32

# Loyalty Tier Table
ws4.merge_cells("A3:F3")
ws4["A3"].value     = "LOYALTY TIER BREAKDOWN"
ws4["A3"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws4["A3"].fill      = fill(MID_NAVY)
ws4["A3"].alignment = center()

loyalty_df = q("""
    SELECT
        CASE WHEN loyalty_points>=5000 THEN 'Platinum (5000+)'
             WHEN loyalty_points>=2000 THEN 'Gold (2000–4999)'
             WHEN loyalty_points>=1000 THEN 'Silver (1000–1999)'
             WHEN loyalty_points>=500  THEN 'Bronze (500–999)'
             ELSE 'Standard (<500)' END  AS loyalty_tier,
        COUNT(*)                         AS customers,
        ROUND(AVG(total_spent),2)        AS avg_spent,
        ROUND(SUM(total_spent),2)        AS total_revenue,
        ROUND(AVG(total_purchases),1)    AS avg_purchases,
        ROUND(AVG(days_since_last_purchase),0) AS avg_days_inactive
    FROM customers GROUP BY loyalty_tier ORDER BY MIN(loyalty_points) DESC
""")
lr1 = write_df(ws4, loyalty_df, start_row=4, start_col=1)

# At-Risk Customers
r2 = lr1 + 3
ws4.merge_cells(f"A{r2}:J{r2}")
ws4[f"A{r2}"].value     = "AT-RISK CUSTOMERS (Inactive > 180 Days)"
ws4[f"A{r2}"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws4[f"A{r2}"].fill      = fill(ACCENT_RED)
ws4[f"A{r2}"].alignment = center()

at_risk = q("""
    SELECT customer_id, full_name, city, state,
           ROUND(total_spent,2) AS total_spent,
           total_purchases, loyalty_points,
           days_since_last_purchase,
           customer_tier,
           CASE WHEN total_spent>=2000 THEN 'HIGH VALUE'
                WHEN total_spent>=500  THEN 'MID VALUE'
                ELSE 'LOW VALUE' END AS risk_priority
    FROM customers
    WHERE days_since_last_purchase > 180 AND total_purchases >= 5
    ORDER BY total_spent DESC LIMIT 30
""")
write_df(ws4, at_risk, start_row=r2+1, start_col=1, header_bg=ACCENT_RED)

# Highlight HIGH VALUE at-risk
for r in range(r2+2, r2+2+len(at_risk)):
    cell = ws4.cell(row=r, column=10)
    if cell.value == "HIGH VALUE":
        for c in range(1, 11):
            ws4.cell(row=r, column=c).fill = fill("FFEBEE")

set_col_width(ws4, {"A":12,"B":22,"C":16,"D":6,"E":14,"F":12,
                    "G":14,"H":20,"I":14,"J":14})

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 — RFM ANALYSIS
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📈 RFM Analysis")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:K1")
ws5["A1"].value     = "RFM (RECENCY · FREQUENCY · MONETARY) SEGMENTATION"
ws5["A1"].font      = Font(name="Arial", size=14, bold=True, color=WHITE)
ws5["A1"].fill      = fill(DARK_NAVY)
ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 32

rfm_full = q("""
    WITH s AS (
        SELECT *,
            CASE WHEN days_since_last_purchase<=30 THEN 5 WHEN days_since_last_purchase<=90 THEN 4
                 WHEN days_since_last_purchase<=180 THEN 3 WHEN days_since_last_purchase<=365 THEN 2 ELSE 1 END AS R,
            CASE WHEN total_purchases>=100 THEN 5 WHEN total_purchases>=50 THEN 4
                 WHEN total_purchases>=25  THEN 3 WHEN total_purchases>=10 THEN 2 ELSE 1 END AS F,
            CASE WHEN total_spent>=10000 THEN 5 WHEN total_spent>=5000 THEN 4
                 WHEN total_spent>=2000  THEN 3 WHEN total_spent>=500  THEN 2 ELSE 1 END AS M
        FROM customers
    )
    SELECT customer_id, full_name, city, state,
           ROUND(total_spent,2) AS total_spent,
           total_purchases,
           days_since_last_purchase,
           R AS recency_score, F AS frequency_score, M AS monetary_score,
           (R+F+M) AS rfm_total,
           CASE WHEN (R+F+M)>=13 THEN 'Champions'
                WHEN (R+F+M)>=10 THEN 'Loyal Customers'
                WHEN (R+F+M)>=8  THEN 'Potential Loyalists'
                WHEN (R+F+M)>=6  THEN 'At Risk'
                WHEN (R+F+M)>=4  THEN 'Hibernating'
                ELSE 'Lost' END AS rfm_segment,
           customer_tier
    FROM s ORDER BY rfm_total DESC
""")
write_df(ws5, rfm_full, start_row=3, start_col=1)

# RFM legend
legend_start = len(rfm_full) + 5
ws5.merge_cells(f"A{legend_start}:D{legend_start}")
ws5[f"A{legend_start}"].value     = "RFM SCORING GUIDE"
ws5[f"A{legend_start}"].font      = Font(name="Arial", size=11, bold=True, color=WHITE)
ws5[f"A{legend_start}"].fill      = fill(MID_NAVY)
ws5[f"A{legend_start}"].alignment = center()

guide = [
    ("Champions",         "13–15", "Top customers: bought recently, buy often, spend most"),
    ("Loyal Customers",   "10–12", "Buy regularly with above-average spending"),
    ("Potential Loyalists","8–9",  "Recent buyers with average frequency"),
    ("At Risk",           "6–7",  "Past good customers, now declining activity"),
    ("Hibernating",       "4–5",  "Last purchase was long ago; low engagement"),
    ("Lost",              "3",    "Lowest recency, frequency and monetary scores"),
]
ws5.cell(row=legend_start+1, column=1, value="Segment").font = Font(bold=True)
ws5.cell(row=legend_start+1, column=2, value="Score Range").font = Font(bold=True)
ws5.cell(row=legend_start+1, column=3, value="Description").font = Font(bold=True)
for ri, (seg, score, desc) in enumerate(guide, start=legend_start+2):
    ws5.cell(row=ri, column=1, value=seg)
    ws5.cell(row=ri, column=2, value=score)
    ws5.cell(row=ri, column=3, value=desc)

set_col_width(ws5, {"A":12,"B":22,"C":16,"D":6,"E":14,"F":12,
                    "G":20,"H":14,"I":16,"J":16,"K":12,"L":20,"M":14})

# ══════════════════════════════════════════════════════════════════════════
# SHEET 6 — RAW DATA
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("📋 Raw Data")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:P1")
ws6["A1"].value     = "COMPLETE CUSTOMER DATASET (500 Records)"
ws6["A1"].font      = Font(name="Arial", size=13, bold=True, color=WHITE)
ws6["A1"].fill      = fill(DARK_NAVY)
ws6["A1"].alignment = center()
ws6.row_dimensions[1].height = 28

raw_data = df.copy()
raw_data["join_date"]          = raw_data["join_date"].astype(str)
raw_data["last_purchase_date"] = raw_data["last_purchase_date"].astype(str)
write_df(ws6, raw_data, start_row=2, start_col=1)

# Auto-filter
ws6.auto_filter.ref = f"A2:{get_column_letter(len(raw_data.columns))}2"

set_col_width(ws6, {
    "A":12,"B":22,"C":6,"D":8,"E":16,"F":6,"G":12,"H":16,
    "I":14,"J":14,"K":16,"L":14,"I":16,"M":22,"N":14,"O":24,"P":14
})

conn.close()
wb.save(OUTPUT_PATH)
print(f"✅  Excel report saved → {OUTPUT_PATH}")
print(f"    Sheets: {[s.title for s in wb.worksheets]}")
